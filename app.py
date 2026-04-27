import io
import json
import math
import os
import re
import sqlite3
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from base64 import b64decode, b64encode
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Font
from rapidfuzz import fuzz, process

from rules import merged_special_rules  # √ľmumi + restoran qaydalarńĪ

# --- ńįnventar emalńĪ (Streamlit Cloud √ľ√ß√ľn app.py daxilind…ô; ayrńĪca .py faylńĪ lazńĪm deyil) ---
_FERQIN_DEYERI_COL_INDEX_FALLBACK = 9
_INVENTORY_DROP_COL_INDEXES: tuple[int, ...] = (
    0,
    1,
    3,
    5,
    10,
    11,
    14,
    15,
    16,
    20,
)
_SORT_COL_INDEX_AFTER_DROPS = 0
_INV_DB_PATH = os.path.join(os.path.dirname(__file__), "inventory_store.db")
_INV_WEEK_SLOTS = ("inv_week1", "inv_week2", "inv_week3", "inv_week4")


def _inv_remote_cfg() -> Optional[dict]:
    """Optional remote persistence via Supabase REST (deploy-proof storage)."""
    try:
        url = st.secrets.get("SUPABASE_URL", "").strip()
        key = st.secrets.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
        table = st.secrets.get("SUPABASE_INV_TABLE", "inventory_files").strip()
    except Exception:
        return None
    if not url or not key:
        return None
    return {
        "url": url.rstrip("/"),
        "key": key,
        "table": table or "inventory_files",
    }


def _inv_remote_request(
    method: str,
    path: str,
    query: Optional[dict] = None,
    payload: Optional[list | dict] = None,
) -> list | dict | None:
    cfg = _inv_remote_cfg()
    if not cfg:
        return None
    q = ""
    if query:
        q = "?" + urllib.parse.urlencode(query, safe="(),=*.")
    url = f"{cfg['url']}{path}{q}"
    body = None
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url=url, method=method.upper(), data=body)
    req.add_header("apikey", cfg["key"])
    req.add_header("Authorization", f"Bearer {cfg['key']}")
    req.add_header("Content-Type", "application/json")
    if method.upper() in ("POST", "PATCH", "PUT"):
        req.add_header("Prefer", "resolution=merge-duplicates,return=representation")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
            if not raw:
                return {}
            return json.loads(raw.decode("utf-8"))
    except Exception:
        return None


def _inv_db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(_INV_DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn


def _inv_db_init() -> None:
    with _inv_db_connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS inventory_files (
                restaurant TEXT NOT NULL,
                slot_key TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_bytes BLOB NOT NULL,
                uploaded_at TEXT NOT NULL,
                PRIMARY KEY (restaurant, slot_key)
            )
            """
        )


def _inv_store_file(restaurant: str, slot_key: str, file_name: str, file_bytes: bytes) -> None:
    cfg = _inv_remote_cfg()
    if cfg:
        payload = [
            {
                "restaurant": restaurant,
                "slot_key": slot_key,
                "file_name": file_name,
                "file_b64": b64encode(file_bytes).decode("ascii"),
                "uploaded_at": datetime.utcnow().isoformat(),
            }
        ]
        out = _inv_remote_request(
            "POST",
            f"/rest/v1/{cfg['table']}",
            payload=payload,
        )
        if out is not None:
            return
    with _inv_db_connect() as conn:
        conn.execute(
            """
            INSERT INTO inventory_files (restaurant, slot_key, file_name, file_bytes, uploaded_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(restaurant, slot_key) DO UPDATE SET
                file_name = excluded.file_name,
                file_bytes = excluded.file_bytes,
                uploaded_at = excluded.uploaded_at
            """,
            (restaurant, slot_key, file_name, file_bytes, datetime.utcnow().isoformat()),
        )


def _inv_get_file(restaurant: str, slot_key: str) -> Optional[dict]:
    cfg = _inv_remote_cfg()
    if cfg:
        rows = _inv_remote_request(
            "GET",
            f"/rest/v1/{cfg['table']}",
            query={
                "select": "file_name,file_b64,uploaded_at",
                "restaurant": f"eq.{restaurant}",
                "slot_key": f"eq.{slot_key}",
                "limit": "1",
            },
        )
        if isinstance(rows, list) and rows:
            row = rows[0]
            try:
                file_bytes = b64decode(row.get("file_b64", ""))
            except Exception:
                file_bytes = b""
            return {
                "name": row.get("file_name", f"{slot_key}.xlsx"),
                "bytes": file_bytes,
                "uploaded_at": row.get("uploaded_at", ""),
            }
    with _inv_db_connect() as conn:
        row = conn.execute(
            """
            SELECT file_name, file_bytes, uploaded_at
            FROM inventory_files
            WHERE restaurant = ? AND slot_key = ?
            """,
            (restaurant, slot_key),
        ).fetchone()
    if row is None:
        return None
    return {
        "name": row["file_name"],
        "bytes": row["file_bytes"],
        "uploaded_at": row["uploaded_at"],
    }


def _inv_delete_file(restaurant: str, slot_key: str) -> None:
    cfg = _inv_remote_cfg()
    if cfg:
        out = _inv_remote_request(
            "DELETE",
            f"/rest/v1/{cfg['table']}",
            query={
                "restaurant": f"eq.{restaurant}",
                "slot_key": f"eq.{slot_key}",
            },
        )
        if out is not None:
            return
    with _inv_db_connect() as conn:
        conn.execute(
            "DELETE FROM inventory_files WHERE restaurant = ? AND slot_key = ?",
            (restaurant, slot_key),
        )


def _inv_count_saved_weeks(restaurant: str) -> int:
    cfg = _inv_remote_cfg()
    if cfg:
        slot_csv = ",".join(_INV_WEEK_SLOTS)
        rows = _inv_remote_request(
            "GET",
            f"/rest/v1/{cfg['table']}",
            query={
                "select": "slot_key",
                "restaurant": f"eq.{restaurant}",
                "slot_key": f"in.({slot_csv})",
            },
        )
        if isinstance(rows, list):
            return len(rows)
    placeholders = ",".join("?" for _ in _INV_WEEK_SLOTS)
    params = [restaurant, *_INV_WEEK_SLOTS]
    with _inv_db_connect() as conn:
        row = conn.execute(
            f"""
            SELECT COUNT(*) AS c
            FROM inventory_files
            WHERE restaurant = ? AND slot_key IN ({placeholders})
            """,
            params,
        ).fetchone()
    return int(row["c"]) if row else 0


def _inv_drop_columns_by_original_positions(
    df: pd.DataFrame, zero_based_indices: tuple[int, ...]
) -> pd.DataFrame:
    cols = list(df.columns)
    names: list[str] = []
    for i in zero_based_indices:
        if 0 <= i < len(cols):
            names.append(cols[i])
    if not names:
        return df
    return df.drop(columns=list(dict.fromkeys(names)), errors="ignore")


def _inv_find_column(df: pd.DataFrame, *header_variants: str) -> Optional[str]:
    want = {h.strip().casefold() for h in header_variants if h.strip()}
    for c in df.columns:
        if str(c).strip().casefold() in want:
            return str(c).strip()
    return None


def _inv_series_nonempty(s: pd.Series) -> pd.Series:
    t = s.map(lambda x: "" if pd.isna(x) else str(x).strip())
    return (t != "") & (t.str.casefold() != "nan") & (t.str.casefold() != "none")


def _inv_parse_decimal(val) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return math.nan
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace("\u00a0", " ")
    if not s or s.lower() in ("nan", "none", "-", "‚ÄĒ"):
        return math.nan
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return math.nan


def _inv_resolve_ferqin_deyeri_column(df: pd.DataFrame) -> Optional[str]:
    col = _inv_find_column(df, "F…ôrqin d…ôy…ôri", "Ferqin deyeri")
    if col is not None:
        return col
    if len(df.columns) > _FERQIN_DEYERI_COL_INDEX_FALLBACK:
        return str(df.columns[_FERQIN_DEYERI_COL_INDEX_FALLBACK]).strip()
    return None


@dataclass(frozen=True)
class InventoryFilterOptions:
    drop_empty_kateqoriya: bool = True
    drop_empty_mahsul: bool = True
    exclude_farqin_open_interval_neg10_pos10: bool = True


def process_inventory_categorization_step(
    orig_xlsx: bytes,
) -> Tuple[Optional[bytes], Optional[str]]:
    try:
        buf = io.BytesIO(orig_xlsx)
        xlf = pd.ExcelFile(buf, engine="openpyxl")
        sheet = xlf.sheet_names[0]
        df = pd.read_excel(xlf, sheet_name=sheet, engine="openpyxl")
    except Exception as e:
        return None, f"Excel oxunmadńĪ: {e}"

    if df.empty:
        return None, "C…ôdv…ôl boŇüdur."

    df = _inv_drop_columns_by_original_positions(df, _INVENTORY_DROP_COL_INDEXES)

    cols = list(df.columns)
    if len(cols) <= _SORT_COL_INDEX_AFTER_DROPS:
        preview = ", ".join(map(str, cols[:25]))
        return None, f"SńĪralama √ľ√ß√ľn A s√ľtunu yoxdur. Qalan s√ľtunlar: {preview}"

    sort_col = cols[_SORT_COL_INDEX_AFTER_DROPS]

    df = df.copy()
    df["__inv_sort_k"] = df[sort_col].map(
        lambda x: str(x).strip().casefold() if pd.notna(x) else "\uffff"
    )
    df = df.sort_values("__inv_sort_k", kind="mergesort").drop(
        columns=["__inv_sort_k"]
    )

    safe_sheet = str(sheet)[:31] if sheet else "Sheet1"
    out = io.BytesIO()
    try:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
    except Exception as e:
        return None, f"Excel yazńĪlmadńĪ: {e}"
    return out.getvalue(), None


def process_inventory_filter_step(
    kateqoriya_xlsx: bytes,
    options: Optional[InventoryFilterOptions] = None,
) -> Tuple[Optional[bytes], Optional[str]]:
    opts = options or InventoryFilterOptions()
    try:
        buf = io.BytesIO(kateqoriya_xlsx)
        xlf = pd.ExcelFile(buf, engine="openpyxl")
        sheet = xlf.sheet_names[0]
        df = pd.read_excel(xlf, sheet_name=sheet, engine="openpyxl")
    except Exception as e:
        return None, f"Excel oxunmadńĪ: {e}"

    if df.empty:
        return None, "C…ôdv…ôl boŇüdur."

    mask = pd.Series(True, index=df.index)

    if opts.drop_empty_kateqoriya:
        col_k = _inv_find_column(df, "Kateqoriya")
        if col_k is not None:
            mask &= _inv_series_nonempty(df[col_k])

    if opts.drop_empty_mahsul:
        col_m = _inv_find_column(df, "M…ôhsul", "Mehsul")
        if col_m is not None:
            mask &= _inv_series_nonempty(df[col_m])

    if opts.exclude_farqin_open_interval_neg10_pos10:
        col_fd = _inv_resolve_ferqin_deyeri_column(df)
        if col_fd is not None:
            vals = df[col_fd].map(_inv_parse_decimal)
            hide = vals.notna() & (vals > -10.0) & (vals < 10.0)
            mask &= ~hide

    df = df.loc[mask].reset_index(drop=True)

    if df.empty:
        return None, "Filtr sonrasńĪ he√ß bir s…ôtir qalmayńĪb."

    safe_sheet = str(sheet)[:31] if sheet else "Sheet1"
    out = io.BytesIO()
    try:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
    except Exception as e:
        return None, f"Excel yazńĪlmadńĪ: {e}"
    return out.getvalue(), None


def process_inventory_emal_pipeline(
    orig_xlsx: bytes,
    filter_options: Optional[InventoryFilterOptions] = None,
) -> Tuple[Optional[bytes], Optional[str]]:
    proc, err = process_inventory_categorization_step(orig_xlsx)
    if err:
        return None, err
    return process_inventory_filter_step(proc, options=filter_options)


st.set_page_config(page_title="ROOM CLOPOS Online", layout="wide")
_inv_db_init()

if "selected_res" not in st.session_state:
    st.session_state.selected_res = "ROOM"
if "last_export" not in st.session_state:
    st.session_state.last_export = None


def _nfc(s: str) -> str:
    s = unicodedata.normalize("NFKC", str(s)).replace("\u00a0", " ")
    s = s.translate(dict.fromkeys(map(ord, "\u200b\u200c\u200d\ufeff"), None))
    return unicodedata.normalize("NFC", s)


def _strip_unicode_marks(s: str) -> str:
    """Ros√©/Rose, g√∂r√ľnm…ôz f…ôrql…ôr ‚ÄĒ Exceld…ôn g…ôl…ôn latńĪn aksentl…ôrini √ßńĪxarńĪr."""
    return "".join(
        ch
        for ch in unicodedata.normalize("NFD", s)
        if unicodedata.category(ch) != "Mn"
    )


def _clean_ad_choices(df_base) -> list:
    """BoŇü / nan s…ôtirl…ôr rapidfuzz-u pozur; bazadan yalnńĪz etibarlńĪ Ad siyahńĪsńĪ."""
    out = []
    for x in df_base["ad"].tolist():
        s = str(x).strip()
        if s and s.lower() not in ("nan", "none"):
            out.append(s)
    return out


def normalize_text(text):
    if not text:
        return ""
    text = _nfc(text).lower().strip()
    text = _strip_unicode_marks(text)
    text = re.sub(r"\(\s*(?:ed|kg|kq|lt|qr|gr|ml|l)\s*\)", "", text)
    text = re.sub(r"\d+\s*%", "", text)
    text = re.sub(r"\d+[\.,]\d+", "", text)
    text = re.sub(r"\b\d+\b", "", text)
    text = re.sub(r"[^\w\s]", " ", text)
    text = (
        text.replace("ne√ľ", "new")
        .replace("c", "k")
        .replace("w", "v")
        .replace("x", "ks")
    )
    text = (
        text.replace("√ß", "c")
        .replace("…ô", "e")
        .replace("ńü", "g")
        .replace("\u0131", "i")
        .replace("i\u0307", "i")
        .replace("√∂", "o")
        .replace("Ňü", "s")
        .replace("√ľ", "u")
    )
    return " ".join(text.split())


def normalize_text_loose(text):
    """R…ôq…ôml…ôri silmir ‚ÄĒ SKU/kod tipli adlar √ľ√ß√ľn; √ßek il…ô baza f…ôrqli olanda …ôsas xilaskar."""
    if not text:
        return ""
    text = _nfc(text).lower().strip()
    text = _strip_unicode_marks(text)
    text = re.sub(r"\(\s*(?:ed|kg|kq|lt|qr|gr|ml|l)\s*\)", "", text)
    text = re.sub(r"\d+\s*%", "", text)
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    text = (
        text.replace("√ß", "c")
        .replace("…ô", "e")
        .replace("ńü", "g")
        .replace("\u0131", "i")
        .replace("√∂", "o")
        .replace("Ňü", "s")
        .replace("√ľ", "u")
        .replace("i\u0307", "i")
    )
    return " ".join(text.split())


def _all_rule_key_tokens_in_receipt(ks: str, n_strict: str) -> bool:
    """√áoxs√∂zl√ľ qayda a√ßarńĪ √ľ√ß√ľn: h…ôr bir uzun k…ôlm…ô √ßekd…ô ayrńĪca token kimi olmalńĪdńĪr.
    ∆Źks halda token_set_ratio m…ôs. yalnńĪz ¬ęananas¬Ľ il…ô ¬ęsandora ananas¬ĽńĪ s…ôhv birl…ôŇüdirir."""
    if not ks or not n_strict:
        return False
    receipt_tokens = set(n_strict.split())
    key_words = [w for w in ks.split() if len(w) > 2]
    if not key_words:
        return True
    for w in key_words:
        if w not in receipt_tokens:
            return False
    return True


def apply_special_logic(name, qty, restaurant: str):
    """rules.py a√ßarńĪ √ßek adńĪnńĪn i√ßind…ô (normallaŇüdńĪrńĪlmńĪŇü) axtarńĪr.
    1) sńĪx + loose alt-s…ôtir; 2) token_set y√ľks…ôk olduqda (f…ôrqli yazńĪlńĪŇü).
    Qaydalar: merged_special_rules(restaurant) ‚ÄĒ COMMON + yalnńĪz h…ômin restoran."""
    if not name or not str(name).strip():
        return name, qty, 1
    raw = str(name).strip()
    n_strict = normalize_text(raw)
    n_loose = normalize_text_loose(raw)
    rules = merged_special_rules(restaurant)
    # ∆Źvv…ôl uzun a√ßar (m…ôs. ¬ęzire zeytun yagi¬Ľ) ‚ÄĒ qńĪsa alt-s…ôtir t…ôsad√ľfi tutulmasńĪn
    rule_items = sorted(
        rules.items(),
        key=lambda kv: len(normalize_text(str(kv[0]))),
        reverse=True,
    )

    for key, val in rule_items:
        ks = normalize_text(str(key))
        kl = normalize_text_loose(str(key))
        if ks and (ks in n_strict or ks in n_loose):
            return str(val[0]).strip(), qty * val[1], val[1]
        if kl and (kl in n_loose or kl in n_strict):
            return str(val[0]).strip(), qty * val[1], val[1]

    for key, val in rule_items:
        ks = normalize_text(str(key))
        if len(ks) < 3:
            continue
        if fuzz.token_set_ratio(ks, n_strict) < 88:
            continue
        # 2+ k…ôlm…ôli a√ßar: b√ľt√ľn k…ôlm…ôl…ôr √ßekd…ô token kimi olmalńĪ (yanlńĪŇü Juice tutulmasńĪnńĪn qarŇüńĪsńĪ)
        if len(ks.split()) >= 2 and not _all_rule_key_tokens_in_receipt(ks, n_strict):
            continue
        rk_sig = _extract_volume_signatures(str(key))
        rec_sig = _extract_volume_signatures(raw)
        if rk_sig and rk_sig != rec_sig:
            continue
        return str(val[0]).strip(), qty * val[1], val[1]

    return name, qty, 1


def _fuzz_proc(x):
    return normalize_text(str(x))


def _fuzz_loose(x):
    return normalize_text_loose(str(x))


def _extract_volume_signatures(text: str) -> frozenset[str]:
    """√áek/baza xam m…ôtnind…ôn h…ôcm imzalarńĪ (l, oz). normalize_text r…ôq…ôml…ôri sildiyi √ľ√ß√ľn
    uyńüunluq qapńĪsńĪ xam s…ôtird…ôn oxuyur ‚ÄĒ 0,33l vs 0,75l, 8 oz vs 12 oz, 2l."""
    if not text:
        return frozenset()
    s = _nfc(str(text)).lower().replace(",", ".")
    out: set[str] = set()
    for m in re.finditer(r"(\d+\.\d+)\s*l\b", s):
        out.add(f"{float(m.group(1))}l")
    for m in re.finditer(r"(?<![\d.])(\d+)\s*l\b", s):
        out.add(f"{float(int(m.group(1)))}l")
    for m in re.finditer(r"(\d+)\s*oz\b", s):
        out.add(f"{int(m.group(1))}oz")
    return frozenset(out)


def _volume_pack_signature_gate(q_raw: str, m_raw: str) -> bool:
    """H…ôcm f…ôrqi olan bazaya s…ôhv birl…ôŇüm…ônin qarŇüńĪsńĪ (Sirab 0,33 vs 0,75, oz)."""
    ql = _nfc(str(q_raw)).lower()
    aq = _extract_volume_signatures(q_raw)
    am = _extract_volume_signatures(m_raw)
    if aq and am:
        return aq == am
    if aq and not am:
        return False
    if not aq and am:
        if "cola" in ql or "sprite" in ql:
            return False
        if "sirab" in ql and "premium" in ql:
            return False
    return True


def _soft_word_gate(q_norm, m_norm, score, strict=False):
    q_words = [w for w in q_norm.split() if len(w) > 2]
    high = 82 if strict else 76
    pr_min = 62 if strict else 52
    if not q_words or score >= high:
        return True
    if any(w in m_norm for w in q_words):
        return True
    return fuzz.partial_ratio(q_norm, m_norm) >= pr_min


def _bar_drink_packaging_gate(q_norm: str, m_norm: str) -> bool:
    """Bar i√ßkil…ôri: Cola vs Cola 2l, Sirab Premium vs Sirab Qazli token_set qarńĪŇüńĪqlńĪńüńĪ."""
    if not q_norm or not m_norm:
        return True
    qc = q_norm.replace(" ", "")
    mc = m_norm.replace(" ", "")
    for needle in ("premium", "zero"):
        if needle in q_norm and needle not in m_norm:
            return False
    for needle in ("2l", "19l"):
        if needle in qc and needle not in mc:
            return False
    if "sirab" in q_norm and "premium" not in q_norm and "premium" in m_norm:
        return False
    if ("cola" in q_norm or "sprite" in q_norm) and "2l" not in qc and "2l" in mc:
        return False
    if ("cola" in q_norm or "sprite" in q_norm) and "2l" in qc and "2l" not in mc:
        return False
    if "sirab" in q_norm:
        q_li = "qazli" in q_norm
        q_siz = "qazsiz" in q_norm
        m_li = "qazli" in m_norm
        m_siz = "qazsiz" in m_norm
        if q_li and (not q_siz) and m_siz and (not m_li):
            return False
        if q_siz and m_li and (not m_siz):
            return False
    return True


def _bar_and_volume_gate(q_raw: str, q_norm: str, m_raw: str, m_norm: str) -> bool:
    return _bar_drink_packaging_gate(q_norm, m_norm) and _volume_pack_signature_gate(
        q_raw, m_raw
    )


def _pick_by_volume_signature(q_raw: str, candidate_strings: list) -> str | None:
    """Eyni normalize_text n…ôtic…ôsi olan adlar arasńĪndan √ßek h…ôcmi il…ô t…ôk baza s…ôtri."""
    if not candidate_strings:
        return None
    if len(candidate_strings) == 1:
        return candidate_strings[0]
    qs = _extract_volume_signatures(q_raw)
    if qs:
        ok = [c for c in candidate_strings if _extract_volume_signatures(c) == qs]
        if len(ok) == 1:
            return ok[0]
        return None
    return candidate_strings[0]


def _match_with_processor(
    q_raw,
    choices,
    threshold,
    proc_fn,
    skip_word_gate=False,
    strict_gate=False,
    score_margin=None,
):
    if not choices:
        return None, 0

    q = str(q_raw).strip()
    if not q or q.lower() == "nan":
        return None, 0

    q_norm = proc_fn(q)
    if not q_norm:
        return None, 0

    norm_equal = [ch for ch in choices if proc_fn(ch) == q_norm]
    if norm_equal:
        gated = [
            str(ch)
            for ch in norm_equal
            if _bar_and_volume_gate(q, q_norm, str(ch), proc_fn(ch))
        ]
        if gated:
            picked = _pick_by_volume_signature(q, gated)
            if picked is not None:
                return picked, 100.0

    # SńĪx norm yoxdursa, loose (aksent/r…ôq…ôm saxlanma) il…ô tam uyńüun ‚ÄĒ Ros√© vs Rose
    if proc_fn is _fuzz_proc:
        q_lo = normalize_text_loose(q)
        loose_equal = [ch for ch in choices if normalize_text_loose(str(ch)) == q_lo]
        if loose_equal:
            gated = [
                str(ch)
                for ch in loose_equal
                if _bar_and_volume_gate(q, q_norm, str(ch), _fuzz_proc(str(ch)))
            ]
            if gated:
                picked = _pick_by_volume_signature(q, gated)
                if picked is not None:
                    return picked, 100.0

    # √áox yaxńĪn tam uyńüunluq (m…ôs. bazada ¬ę6 li¬Ľ / √ßekd…ô ¬ę6li¬Ľ) ‚ÄĒ ki√ßik f…ôrql…ôri tutur
    if len(choices) <= 4000 and len(q_norm) >= 6:
        ratio_pass: list[tuple[str, float]] = []
        for choice in choices:
            cn = proc_fn(choice)
            r = fuzz.ratio(q_norm, cn)
            if r < 98:
                continue
            if not _bar_and_volume_gate(q, q_norm, str(choice), cn):
                continue
            ratio_pass.append((str(choice), float(r)))
        if ratio_pass:
            max_r = max(x[1] for x in ratio_pass)
            top = [x[0] for x in ratio_pass if x[1] == max_r]
            picked = _pick_by_volume_signature(q, top)
            if picked is not None:
                return picked, float(min(max_r, 99.9))

    # 3+ k…ôlm…ô: WRatio tam ifad…ôni (m…ôs. ¬ęsensoy sweet chili¬Ľ) token_set-d…ôn yaxŇüńĪ tuta bil…ôr
    n_words = len(q_norm.split())
    primary_scorer = fuzz.WRatio if n_words >= 3 else fuzz.token_set_ratio

    best = process.extractOne(
        q,
        choices,
        scorer=primary_scorer,
        processor=proc_fn,
    )
    if not best:
        return None, 0

    best_match = str(best[0])
    score = float(best[1])
    used_alt_scorer = False

    if score < threshold:
        alt_scorer = (
            fuzz.token_set_ratio if primary_scorer == fuzz.WRatio else fuzz.WRatio
        )
        best2 = process.extractOne(q, choices, scorer=alt_scorer, processor=proc_fn)
        if best2 and float(best2[1]) >= threshold:
            best_match = str(best2[0])
            score = float(best2[1])
            used_alt_scorer = True

    m_norm = proc_fn(best_match)
    if not skip_word_gate and not _soft_word_gate(
        q_norm, m_norm, score, strict=strict_gate
    ):
        return None, score

    if score < threshold:
        return None, score

    if not _bar_and_volume_gate(q, q_norm, best_match, m_norm):
        alt_found = None
        alt_score = 0.0
        for cand, sc, _ in process.extract(
            q,
            choices,
            scorer=primary_scorer,
            processor=proc_fn,
            limit=35,
        ):
            scf = float(sc)
            if scf < threshold:
                continue
            cn = proc_fn(cand)
            if not skip_word_gate and not _soft_word_gate(
                q_norm, cn, scf, strict=strict_gate
            ):
                continue
            if not _bar_and_volume_gate(q, q_norm, str(cand), cn):
                continue
            alt_found, alt_score = str(cand), scf
            break
        if alt_found is None:
            return None, score
        best_match, score, m_norm = alt_found, alt_score, proc_fn(alt_found)
        used_alt_scorer = True

    # ńįki baza s…ôtri eyni xala yaxńĪndńĪrsa ‚ÄĒ s…ôhv se√ßim riski; √ßox y√ľks…ôk xalda margin t…ôtbiq olunmur
    if (
        score_margin is not None
        and score < 99.9
        and score < 88
        and not used_alt_scorer
        and not skip_word_gate
    ):
        topn = process.extract(
            q, choices, scorer=primary_scorer, processor=proc_fn, limit=2
        )
        if (
            len(topn) >= 2
            and str(topn[0][0]) == best_match
            and (float(topn[0][1]) - float(topn[1][1])) < score_margin
        ):
            return None, float(topn[0][1])

    return best_match, score


def get_best_match(
    query_name, choices, threshold=74, *, safe_mode=True, score_margin=7.0
):
    """SńĪx ‚Üí loose; t…ôhl√ľk…ôsiz rejimd…ô son √ßar…ô v…ô Ňü√ľbh…ôli ¬ęikinci yer¬Ľ yaxńĪnlńĪńüńĪ s√∂nd√ľr√ľl√ľr."""
    sm = float(score_margin) if (safe_mode and score_margin is not None) else None
    sg = bool(safe_mode)
    r = _match_with_processor(
        query_name, choices, threshold, _fuzz_proc, strict_gate=sg, score_margin=sm
    )
    if r[0]:
        return r
    loose_thr = max(56, int(threshold) - 6)
    r = _match_with_processor(
        query_name,
        choices,
        loose_thr,
        _fuzz_loose,
        strict_gate=False,
        score_margin=sm,
    )
    if r[0]:
        return r
    if safe_mode:
        return None, 0
    qn = _fuzz_proc(query_name)
    if len(qn) < 5 or len(qn.split()) < 2:
        return None, 0
    last_thr = max(40, int(threshold) - 22)
    return _match_with_processor(
        query_name,
        choices,
        last_thr,
        _fuzz_loose,
        skip_word_gate=True,
        strict_gate=False,
        score_margin=None,
    )


def explain_match(query_name, choices, limit=5, processor=None):
    q = str(query_name).strip()
    if not choices or not q:
        return []
    proc = processor if processor is not None else _fuzz_proc
    return process.extract(
        q,
        choices,
        scorer=fuzz.token_set_ratio,
        processor=proc,
        limit=limit,
    )


def parse_az_number(val):
    """Excel AZ formatńĪ: verg√ľl onluq (1,135), boŇüluqlu minlik nadir."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace("\u00a0", " ")
    if not s or s.lower() in ("nan", "none", "-", "‚ÄĒ"):
        return 0.0
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def standardize_columns(df, chek_fayli=False):
    """chek_fayli=True: sklad √ßekind…ôki s…ôtir ¬ęID¬Ľ s√ľtunu bazanńĪn Clopos id-si il…ô qarńĪŇümasńĪn."""
    df = df.copy()
    df.columns = [str(c).strip().lstrip("\ufeff") for c in df.columns]
    renamed = {}
    for col in df.columns:
        key = normalize_text(col)
        col_l = str(col).lower()
        if key == "ad" or "mehsul" in key or "nomenkl" in key or key in ("mal", "title", "name"):
            renamed[col] = "ad"
        elif (
            "miqdar" in key
            or "kemiyyat" in key
            or key in ("say", "qty", "qty.")
            or "eded" in key
        ):
            renamed[col] = "miqdar"
        elif "umumi" in key or ("maya" in key and "dey" in key):
            renamed[col] = "line_total_src"
        elif key == "vahid" and "‚āľ" not in col_l and "azn" not in col_l:
            # YalnńĪz √∂l√ß√ľ vahidi (kg/pcs) ‚ÄĒ qiym…ôt deyil
            renamed[col] = "unit_kind"
        elif "vahid" in key and ("‚āľ" in col_l or "azn" in col_l or "qiym" in key):
            renamed[col] = "price"
        elif any(k in key for k in ["qiym", "azn"]) or "‚āľ" in col_l:
            renamed[col] = "price"
        elif key == "id":
            renamed[col] = "cek_line_id" if chek_fayli else "id"
    return df.rename(columns=renamed)


def normalize_restaurant_name(name):
    return str(name).lower().replace("\u0131", "i").replace("i\u0307", "i").strip()


def discover_restaurants():
    restaurants = set()
    for file_name in os.listdir("."):
        lower_name = file_name.lower()
        if not lower_name.startswith("ana_"):
            continue
        if not (lower_name.endswith(".xlsx") or lower_name.endswith(".csv")):
            continue
        if "_horeca" in lower_name:
            restaurants.add(file_name[4:].rsplit("_horeca", 1)[0].upper())
        elif "_dk" in lower_name:
            restaurants.add(file_name[4:].rsplit("_dk", 1)[0].upper())
    return sorted(restaurants) if restaurants else ["ROOM", "BIBLIOTEKA", "FINESTRA"]


def _resolve_db_path_for_suffix(res_name, suffix: str):
    """suffix: 'horeca' v…ô ya 'dk' ‚ÄĒ fayl adńĪ ana_<restoran>_<suffix>."""
    target_prefix = f"ana_{normalize_restaurant_name(res_name)}_{suffix}"
    for file_name in os.listdir("."):
        normalized_file = normalize_restaurant_name(file_name)
        if normalized_file.startswith(target_prefix):
            if file_name.lower().endswith((".xlsx", ".csv")):
                return file_name
    return None


def _read_single_db_path(path):
    if not path:
        return None
    try:
        if path.lower().endswith(".xlsx"):
            return pd.read_excel(path)
        return pd.read_csv(path)
    except Exception:
        return None


@st.cache_data(ttl=30, show_spinner=False)
def get_db(res_name, category):
    """Horeca: yalnńĪz *_horeca. Dark Kitchen: *_dk + *_horeca birl…ôŇüik (eyni Ad t…ôkrarlanmasa, dk √ľst√ľn)."""
    if category == "Horeca":
        raw = _read_single_db_path(_resolve_db_path_for_suffix(res_name, "horeca"))
        if raw is None or raw.empty:
            return None
        return standardize_columns(raw, chek_fayli=False)
    parts = []
    for suffix in ("dk", "horeca"):
        raw = _read_single_db_path(_resolve_db_path_for_suffix(res_name, suffix))
        if raw is None or raw.empty:
            continue
        parts.append(standardize_columns(raw, chek_fayli=False))
    if not parts:
        return None
    out = pd.concat(parts, ignore_index=True)
    # Eyni baŇülńĪqla iki s√ľtun (m…ôs. biri ¬ęAd¬Ľ, biri ¬ęad¬Ľ) olanda df["ad"] DataFrame olur ‚Üí .str x…ôtasńĪ
    out = out.loc[:, ~out.columns.duplicated(keep="first")]
    if "ad" in out.columns:
        out = out.drop_duplicates(subset=["ad"], keep="first")
    return out


def _clean_receipt_no(v: str) -> str:
    s = _nfc(str(v)).strip()
    s = re.sub(r"[^\w\-]+", " ", s, flags=re.UNICODE)
    s = " ".join(s.split())
    return s[:40]


def _extract_export_receipt_no(df_c, uploaded_name: str = "") -> str:
    if isinstance(df_c, pd.DataFrame) and not df_c.empty:
        cols = [str(c) for c in df_c.columns]
        priority = [
            "cek_no",
            "check_no",
            "chek_no",
            "sened_no",
            "nomre",
            "id",
            "cek_line_id",
        ]
        ordered = [c for c in priority if c in cols]
        ordered += [
            c
            for c in cols
            if c not in ordered
            and (
                re.search(r"(cek|check|chek).*(no|nom|nmr)", c, flags=re.IGNORECASE)
                or re.search(r"(no|nom|nmr).*(cek|check|chek)", c, flags=re.IGNORECASE)
                or re.search(r"nomre|sened", c, flags=re.IGNORECASE)
            )
        ]
        for c in ordered:
            s = (
                df_c[c]
                .dropna()
                .astype(str)
                .map(lambda x: _nfc(x).strip())
            )
            s = s[~s.str.lower().isin(["", "nan", "none"])]
            if s.empty:
                continue
            counts = s.value_counts(dropna=True)
            top = str(counts.index[0]).strip()
            if len(counts) == 1 or int(counts.iloc[0]) >= max(2, int(len(s) * 0.6)):
                cleaned = _clean_receipt_no(top)
                if cleaned:
                    return cleaned
    if uploaded_name:
        m = re.search(r"(\d{4,})", _nfc(uploaded_name))
        if m:
            return m.group(1)
    return datetime.now().strftime("%Y%m%d_%H%M")


def build_export_file_name(restaurant, category, receipt_no=None):
    rec = _clean_receipt_no(receipt_no or "")
    if rec:
        return f"clopos import {rec}.xlsx"
    category_tag = "horeca" if category == "Horeca" else "dk"
    restaurant_tag = normalize_restaurant_name(restaurant).replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return f"clopos_{restaurant_tag}_{category_tag}_{timestamp}.xlsx"


def _excel_sheet_no_bold(writer, sheet_name):
    """B√ľt√ľn xanalar normal Ňürift ‚ÄĒ ID/QUANTITY/COST baŇülńĪńüńĪ da qalńĪn olmasńĪn."""
    ws = writer.sheets[sheet_name]
    plain = Font(bold=False)
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.font = plain


def to_bold_excel_bytes(dataframe):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="CLOPOS")
        _excel_sheet_no_bold(writer, "CLOPOS")
    output.seek(0)
    return output.getvalue()


def to_tapilmayan_only_bytes(unmatched_df):
    """YalnńĪz …ôl il…ô iŇü √ľ√ß√ľn TapńĪlmayanlar v…ôr…ôqi."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        unmatched_df.to_excel(writer, index=False, sheet_name="Tapilmayanlar")
        _excel_sheet_no_bold(writer, "Tapilmayanlar")
    output.seek(0)
    return output.getvalue()


def _resolve_id_for_product(df_base, name_query):
    """Bazada Ad il…ô ID tapńĪr: d…ôqiq, b√∂y√ľk/ki√ßik, NFC, normallaŇüdńĪrńĪlmńĪŇü, son √ßar…ô ratio.
    Qaydadan g…ôl…ôn ad Exceld…ô ¬ęeyni g√∂r√ľn…ôn¬Ľ amma Unicode f…ôrqli olanda k√∂m…ôk edir."""
    if name_query is None:
        return None, None
    raw = str(name_query).strip()
    if not raw or raw.lower() in ("nan", "none"):
        return None, None

    ads = df_base["ad"].astype(str).str.strip()

    def _pick(mask):
        if not mask.any():
            return None, None
        row0 = df_base.loc[mask].iloc[0]
        return int(row0["id"]), str(row0["ad"]).strip()

    m, ad = _pick(ads == raw)
    if m is not None:
        return m, ad

    m, ad = _pick(ads.str.lower() == raw.lower())
    if m is not None:
        return m, ad

    raw_nfc = _strip_unicode_marks(_nfc(raw).strip().lower())
    nfc_match = ads.map(
        lambda x: _strip_unicode_marks(_nfc(str(x)).strip().lower())
    ) == raw_nfc
    m, ad = _pick(nfc_match)
    if m is not None:
        return m, ad

    q_nl = normalize_text_loose(raw)
    loose_match = ads.map(lambda x: normalize_text_loose(str(x))) == q_nl
    m, ad = _pick(loose_match)
    if m is not None:
        return m, ad

    qn = normalize_text(raw)
    norm_match = ads.map(lambda x: normalize_text(str(x))) == qn
    m, ad = _pick(norm_match)
    if m is not None:
        return m, ad

    raw_cmp = _strip_unicode_marks(_nfc(raw).strip().lower())
    best_r, best_id, best_ad = 0, None, None
    for _, row in df_base.iterrows():
        ad = str(row["ad"]).strip()
        r = fuzz.ratio(raw_cmp, _strip_unicode_marks(_nfc(ad).strip().lower()))
        if r > best_r:
            best_r = r
            best_id = int(row["id"])
            best_ad = ad
    if best_id is not None and best_r >= 96:
        return best_id, best_ad
    if len(raw_cmp) >= 10:
        best_ts, best_id2, best_ad2 = 0, None, None
        for _, row in df_base.iterrows():
            ad = str(row["ad"]).strip()
            ts = float(
                fuzz.token_set_ratio(
                    raw_cmp, _strip_unicode_marks(_nfc(ad).strip().lower())
                )
            )
            if ts > best_ts:
                best_ts, best_id2, best_ad2 = ts, int(row["id"]), ad
        if best_id2 is not None and best_ts >= 93.0:
            return best_id2, best_ad2
    return None, None


def _first_id_for_name(df_base, m_name):
    """K√∂hn…ô √ßańüńĪrńĪŇülar √ľ√ß√ľn; uńüursuzdursa KeyError."""
    mid, _ = _resolve_id_for_product(df_base, m_name)
    if mid is None:
        raise KeyError(f"id tapńĪlmadńĪ: {m_name!r}")
    return mid


def _render_restoran_online_panel() -> None:
    """Analiz v…ô Kontrol ‚ÄĒ yalnńĪz Restoran Ňü√∂b…ôsi aktiv olanda g√∂st…ôrilir."""
    curr = st.session_state.selected_res
    st.markdown(
        f"<h3 style='text-align: center;'>{curr} | Online Panel</h3>",
        unsafe_allow_html=True,
    )
    tab1, tab2 = st.tabs(["ūüöÄ ANALńįZ", "ūüĒć KONTROL"])
    
    with tab1:
        st.caption(
            "Mexanizm: √ßekd…ôki **Ad** il…ô ana bazada **eyni m…ôhsul adńĪ** tapńĪlńĪr ‚Üí export **ID** "
            "yalnńĪz bazadandńĪr. **QUANTITY** = √ßek miqdarńĪ (x√ľsusi qayda varsa √ßevrilmiŇü miqdar). "
            "**COST** = (√ßekd…ôki vahid qiym…ôt √∑ Miqdar) √∑ **qayda faktoru** (…ôg…ôr 1-dirs…ô yalnńĪz birinci b√∂lm…ô): "
            "m…ôs. 1 paket 5 kq = 7 ‚āľ ‚Üí √ßek vahidi 7 ‚āľ, faktor 5 ‚Üí **1 kq √ľ√ß√ľn 1,4 ‚āľ**; QUANTITY 5 kq olunca s…ôtir c…ômi 7 ‚āľ. "
            "**T…ôhl√ľk…ôsiz rejim**: yalnńĪz aydńĪn uyńüunluq q…ôbul edilir; qalanlar **TapńĪlmayanlar** "
            "v…ôr…ôqind…ô …ôl il…ô doldurmaq √ľ√ß√ľnd√ľr. **Dark Kitchen** √ľ√ß√ľn ana baza h…ôm `ana_<rest>_dk`, "
            "h…ôm `ana_<rest>_horeca` faylńĪndan oxunur v…ô birl…ôŇüdirilir; **Horeca** √ľ√ß√ľn yalnńĪz `_horeca`."
        )
        col_a, col_b, col_c = st.columns([1, 1, 1])
        cat = col_a.selectbox("Sah…ô:", ["Horeca", "Dark Kitchen"])
        aggressive_match = col_a.checkbox(
            "Agressiv uyńüunluq (daha √ßox avtomatik s…ôtir, daha √ßox s…ôhv riski)",
            value=False,
            help="S√∂nd√ľr√ľlm√ľŇüd…ô: y√ľks…ôk h…ôdd, ikinci yer…ô yaxńĪn n…ôtic…ôl…ôr r…ôdd, son √ßar…ô yoxdur.",
        )
        match_thr = col_c.slider(
            "Uyńüunluq h…ôddi (%) ‚ÄĒ t…ôhl√ľk…ôsiz rejimd…ô t…ôklif: 72‚Äď80",
            min_value=60,
            max_value=95,
            value=74,
            help="AŇüańüńĪ = daha √ßox s…ôtir; t…ôhl√ľk…ôsiz rejimd…ô Ňü√ľbh…ôli yaxńĪnlńĪqlar yen…ô r…ôdd edil…ô bil…ôr.",
        )
        cek = col_b.file_uploader("ūüďĄ Sklad √áekini Y√ľkl…ô", type=["xlsx"])
    
        if cek and st.button("‚ö° BaŇülat"):
            df_base = get_db(curr, cat)
            if df_base is not None:
                df_c = pd.read_excel(cek)
                df_c = standardize_columns(df_c, chek_fayli=True)
                df_base = standardize_columns(df_base, chek_fayli=False)
    
                required_cek = {"ad", "miqdar"}
                required_base = {"ad", "id"}
                if not required_cek.issubset(set(df_c.columns)):
                    st.markdown("**Problem:** √áek faylńĪnda `Ad` v…ô `Miqdar` s√ľtunlarńĪ tapńĪlmadńĪ.")
                    st.stop()
                if not required_base.issubset(set(df_base.columns)):
                    st.markdown("**Problem:** Baza faylńĪnda `Ad` v…ô `id` s√ľtunlarńĪ tapńĪlmadńĪ.")
                    st.stop()
    
                # choices strip olunur; df_base["ad"] d…ô eyni olmalńĪdńĪr ‚ÄĒ …ôks halda id tapńĪlmńĪr
                df_c["ad"] = df_c["ad"].astype(str).str.strip()
                for _col in ("miqdar", "price"):
                    if _col in df_c.columns:
                        df_c[_col] = df_c[_col].map(parse_az_number)
                df_base["ad"] = df_base["ad"].astype(str).str.strip()
                df_base["id"] = pd.to_numeric(df_base["id"], errors="coerce")
                df_base = df_base.dropna(subset=["id", "ad"])
                df_base["id"] = df_base["id"].astype(int)
                df_base = df_base.drop_duplicates(subset=["ad"], keep="first")
    
                final_list = []
                errors = 0
                choices = _clean_ad_choices(df_base)
                fail_debug = []
                tapilmayan_rows = []
                skipped_rows = []
                safe_mode = not aggressive_match
                for row_idx, (_, row) in enumerate(df_c.iterrows(), start=1):
                    o_name = ""
                    p_name = ""
                    try:
                        o_name = str(row.get("ad", "")).strip()
                        if not o_name or o_name.lower() in ("nan", "none"):
                            skipped_rows.append(
                                {
                                    "S…ôtir": row_idx,
                                    "√áekd…ôki_ad": str(row.get("ad", ""))[:240],
                                    "S…ôb…ôb": "BoŇü / etibarsńĪz ad",
                                }
                            )
                            continue
                        o_qty = parse_az_number(row.get("miqdar", 0))
                        unit_price = parse_az_number(row.get("price", 0))
                        if o_qty == 0:
                            skipped_rows.append(
                                {
                                    "S…ôtir": row_idx,
                                    "√áekd…ôki_ad": o_name,
                                    "S…ôb…ôb": "Miqdar 0",
                                }
                            )
                            continue
    
                        p_name, p_qty, fct = apply_special_logic(o_name, o_qty, curr)
                        # √áek miqdarńĪna g√∂r…ô bir fiziki vahidin (paket, ed) qiym…ôti; qayda il…ô miqdar fct vurulanda
                        # Clopos COST = baza vahidinin (m…ôs. 1 kq) qiym…ôti ‚Üí paket qiym…ôti / fct (7/5=1.4).
                        price_per_cheque_unit = (unit_price / o_qty) if o_qty != 0 else 0.0
                        cost = (
                            (price_per_cheque_unit / fct)
                            if fct not in (None, 0)
                            else price_per_cheque_unit
                        )
                        m_name, _score = get_best_match(
                            p_name,
                            choices,
                            threshold=match_thr,
                            safe_mode=safe_mode,
                        )
                        mid, _canon = (None, None)
                        if m_name:
                            mid, _canon = _resolve_id_for_product(df_base, m_name)
                        if mid is None:
                            mid, _canon = _resolve_id_for_product(df_base, p_name)
                        if mid is not None:
                            final_list.append(
                                {
                                    "ID": mid,
                                    "QUANTITY": p_qty,
                                    "COST": round(cost, 4),
                                    "LINE_TOTAL": round(p_qty * cost, 4),
                                }
                            )
                        else:
                            errors += 1
                            hits = explain_match(p_name, choices, limit=5)
                            hits_l = explain_match(
                                p_name, choices, limit=3, processor=_fuzz_loose
                            )
                            row_dbg = {
                                "√áekd…ô ad": o_name,
                                "Miqdar": o_qty,
                                "Bir_vahid_COST": round(cost, 4),
                                "Qaydadan sonra": p_name,
                                "∆Źn yaxńĪn (token_set)": hits[0][0] if hits else "",
                                "Xal": round(float(hits[0][1]), 1) if hits else "",
                                "2-ci": hits[1][0] if len(hits) > 1 else "",
                                "2 xal": round(float(hits[1][1]), 1) if len(hits) > 1 else "",
                                "Loose 1": hits_l[0][0] if hits_l else "",
                                "Loose xal": round(float(hits_l[0][1]), 1) if hits_l else "",
                            }
                            fail_debug.append(row_dbg)
                            tapilmayan_rows.append(
                                {
                                    "S…ôtir": row_idx,
                                    "√áekd…ôki_ad": o_name,
                                    "Qaydadan_sonra": p_name,
                                    "Miqdar": o_qty,
                                    "Bir_vahid_COST": round(cost, 4),
                                    "ID_…ôl_ile": "",
                                    "Baza_AdńĪ_…ôl_ile": "",
                                }
                            )
                    except (ValueError, TypeError, KeyError) as ex:
                        errors += 1
                        eq_x = parse_az_number(row.get("miqdar", 0))
                        up_x = parse_az_number(row.get("price", 0))
                        cst_x = (up_x / eq_x) if eq_x else 0.0
                        fail_debug.append(
                            {
                                "√áekd…ô ad": o_name,
                                "Miqdar": eq_x,
                                "Bir_vahid_COST": round(cst_x, 4) if eq_x else "",
                                "Qaydadan sonra": p_name,
                                "∆Źn yaxńĪn (token_set)": f"(x…ôta) {type(ex).__name__}",
                                "Xal": "",
                                "2-ci": str(ex)[:120],
                                "2 xal": "",
                            }
                        )
                        tapilmayan_rows.append(
                            {
                                "S…ôtir": row_idx,
                                "√áekd…ôki_ad": o_name
                                or str(row.get("ad", "")).strip()
                                or "(x…ôta)",
                                "Qaydadan_sonra": p_name,
                                "Miqdar": eq_x if eq_x else "",
                                "Bir_vahid_COST": round(cst_x, 4) if eq_x else "",
                                "ID_…ôl_ile": "",
                                "Baza_AdńĪ_…ôl_ile": "",
                            }
                        )
                        continue
    
                n_cek = len(df_c)
                n_tutulan = len(final_list)
                n_tapilmayan = len(tapilmayan_rows)
                n_kecilen = len(skipped_rows)
                st.info(
                    f"**√áek s…ôtri (c…ômi):** {n_cek} | **Avtomatik tutulan:** {n_tutulan} | "
                    f"**TapńĪlmayan:** {n_tapilmayan} | **Ke√ßil…ôn (boŇü ad / miqdar 0):** {n_kecilen}  \n"
                    f"*(Yoxlama: {n_tutulan} + {n_tapilmayan} + {n_kecilen} = {n_tutulan + n_tapilmayan + n_kecilen} ‚ÄĒ √ßekl…ô uyńüun g…ôlm…ôlidir.)*"
                )
                if skipped_rows:
                    with st.expander("Ke√ßil…ôn s…ôtirl…ôr (boŇü ad v…ô ya miqdar 0)", expanded=False):
                        st.dataframe(pd.DataFrame(skipped_rows), use_container_width=True)
    
                if not final_list:
                    st.markdown(
                        "**Diqq…ôt:** Uyńüun m…ôhsul tapńĪlmadńĪ. Ad yazńĪlńĪŇülarńĪ f…ôrqli ola bil…ôr v…ô ya baza faylńĪ uyńüun deyil."
                    )
                    st.info(
                        f"Baza m…ôhsulu: {len(df_base)} | Uńüursuz match c…ôhdi: {errors}"
                    )
                    if tapilmayan_rows:
                        st.markdown("### TapńĪlmayanlar")
                        st.caption(
                            "Bu s…ôtirl…ôri …ôl il…ô bazada tapńĪb `rules.py` v…ô ya ana bazaya …ôlav…ô edin; "
                            "boŇü s√ľtunlarńĪ Exceld…ô doldura bil…ôrsiniz."
                        )
                        um_only = pd.DataFrame(tapilmayan_rows)
                        st.dataframe(um_only, use_container_width=True)
                        st.download_button(
                            "ūüď• TapńĪlmayanlar (Excel)",
                            to_tapilmayan_only_bytes(um_only),
                            f"tapilmayanlar_{curr}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            key="download_um_only_fail",
                        )
                    sample = df_c[["ad", "miqdar"]].dropna(subset=["ad"]).head(10).rename(
                        columns={"ad": "√áekd…ô ad", "miqdar": "Miqdar"}
                    )
                    if not sample.empty:
                        st.markdown("ńįlk 10 √ßek adńĪ (baza il…ô vizual m√ľqayis…ô √ľ√ß√ľn):")
                        st.dataframe(sample, use_container_width=True)
                    if fail_debug:
                        dbg_df = pd.DataFrame(fail_debug)
                        with st.expander(
                            "Diaqnostika: h…ôr s…ôtir √ľ√ß√ľn bazadan …ôn yaxńĪn 2 variant (xal aŇüańüńĪdńĪrsa h…ôddi sal)",
                            expanded=True,
                        ):
                            st.dataframe(dbg_df, use_container_width=True)
                        dbg_bytes = to_bold_excel_bytes(
                            dbg_df.rename(
                                columns={
                                    "√áekd…ô ad": "cek_ad",
                                    "Qaydadan sonra": "qayda_sonra",
                                    "∆Źn yaxńĪn (token_set)": "en_yaxin_1",
                                    "Xal": "xal_1",
                                    "2-ci": "en_yaxin_2",
                                    "2 xal": "xal_2",
                                    "Loose 1": "loose_1",
                                    "Loose xal": "loose_xal",
                                }
                            )
                        )
                        st.download_button(
                            "ūüď• Diaqnostika c…ôdv…ôlini Excel kimi endir",
                            dbg_bytes,
                            f"clopos_diag_{curr}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            key="download_diag",
                        )
                    st.caption(
                        "∆Źsas export yalnńĪz …ôn azńĪ bir s…ôtir uńüurla uyńüunlaŇüanda √ßńĪxńĪr. "
                        "YuxarńĪdakńĪ s√ľrg√ľ il…ô h…ôddi azaldńĪb ‚ö° BaŇülat-a yenid…ôn bas."
                    )
                    st.stop()
    
                res_df = (
                    pd.DataFrame(final_list)
                    .groupby("ID", as_index=False)
                    .agg({"QUANTITY": "sum", "LINE_TOTAL": "sum"})
                )
                res_df["COST"] = (res_df["LINE_TOTAL"] / res_df["QUANTITY"]).round(4)
                res_df = res_df[["ID", "QUANTITY", "COST"]]
    
                st.markdown(f"**HazńĪrdńĪr:** {len(res_df)} m…ôhsul hazńĪrlandńĪ.")
                if errors:
                    st.info(
                        f"{errors} s…ôtir avtomatik tutulmadńĪ ‚ÄĒ **TapńĪlmayanlar** b√∂lm…ôsind…ô …ôl il…ô iŇül…ôyin."
                    )
    
                st.dataframe(res_df, use_container_width=True)
    
                um_df = pd.DataFrame(tapilmayan_rows) if tapilmayan_rows else pd.DataFrame()
                if not um_df.empty:
                    st.markdown("### TapńĪlmayanlar")
                    st.caption(
                        "Bu s…ôtirl…ôr t…ôhl√ľk…ôsiz rejimd…ô baza il…ô avtomatik birl…ôŇüdirilm…ôdi (v…ô ya uyńüunluq "
                        "Ňü√ľbh…ôli sayńĪldńĪ). **ID_…ôl_ile** / **Baza_AdńĪ_…ôl_ile** s√ľtunlarńĪnńĪ Exceld…ô doldurub "
                        "sonra `rules.py`-…ô qayda …ôlav…ô edin v…ô ya bazanńĪ yenil…ôyin."
                    )
                    st.dataframe(um_df, use_container_width=True)
                    st.download_button(
                        "ūüď• YalnńĪz TapńĪlmayanlar (Excel)",
                        to_tapilmayan_only_bytes(um_df),
                        f"tapilmayanlar_{curr}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        key="download_um_only_ok",
                    )
    
                receipt_no = _extract_export_receipt_no(df_c, getattr(cek, "name", ""))
                export_name = build_export_file_name(curr, cat, receipt_no=receipt_no)
                export_bytes = to_bold_excel_bytes(res_df)
                um_bytes = (
                    to_tapilmayan_only_bytes(um_df) if not um_df.empty else None
                )
                um_name = (
                    f"tapilmayanlar_{curr}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    if um_bytes
                    else None
                )
                st.session_state.last_export = {
                    "restaurant": curr,
                    "category": cat,
                    "rows": len(res_df),
                    "unmatched": len(um_df),
                    "file_name": export_name,
                    "file_bytes": export_bytes,
                    "unmatched_bytes": um_bytes,
                    "unmatched_file_name": um_name,
                    "preview_df": res_df,
                }
                st.markdown(
                    "**Clopos import:** `"
                    + export_name
                    + "`"
                    + (
                        f" ‚ÄĒ …ôlav…ô olaraq **TapńĪlmayanlar** √ľ√ß√ľn ayrńĪ Excel ({len(um_df)} s…ôtir) endir."
                        if not um_df.empty
                        else "."
                    )
                )
                st.download_button(
                    "ūüď• Clopos import (yalnńĪz ID, QUANTITY, COST)",
                    export_bytes,
                    export_name,
                    key="download_current",
                )
            else:
                dk_hint = (
                    " **Dark Kitchen** √ľ√ß√ľn …ôn azńĪ biri olmalńĪdńĪr: `ana_<restoran>_dk` v…ô/v…ô ya "
                    "`ana_<restoran>_horeca`."
                    if cat == "Dark Kitchen"
                    else " **Horeca** √ľ√ß√ľn `ana_<restoran>_horeca` faylńĪ lazńĪmdńĪr."
                )
                st.markdown(
                    "**Problem:** Uyńüun ana baza tapńĪlmadńĪ. Repo k√∂k√ľnd…ô fayl adńĪ `ana_<restoran>_horeca` "
                    "v…ô ya `ana_<restoran>_dk` formatńĪnda olmalńĪdńĪr."
                    + dk_hint
                )
    
        saved_export = st.session_state.get("last_export")
        if saved_export:
            st.markdown("---")
            st.markdown("### Son hazńĪrlanmńĪŇü fayl")
            st.write(
                f"Restoran: **{saved_export['restaurant']}** | "
                f"Sah…ô: **{saved_export['category']}** | "
                f"S…ôtir sayńĪ: **{saved_export['rows']}**"
            )
            if saved_export.get("unmatched"):
                st.caption(
                    f"Son analizd…ô **tapńĪlmayan** s…ôtir: **{saved_export['unmatched']}** ‚ÄĒ Clopos faylńĪna "
                    "daxil edilmir; ayrńĪca Excel aŇüańüńĪdan endir."
                )
            st.write(f"Clopos faylńĪ: `{saved_export['file_name']}`")
            st.dataframe(saved_export["preview_df"], use_container_width=True)
            st.download_button(
                "ūüď• Clopos faylńĪnńĪ yenid…ôn endir",
                saved_export["file_bytes"],
                saved_export["file_name"],
                key="download_saved",
            )
            ub = saved_export.get("unmatched_bytes")
            ufn = saved_export.get("unmatched_file_name")
            if ub and ufn:
                st.download_button(
                    "ūüď• TapńĪlmayanlar faylńĪnńĪ yenid…ôn endir",
                    ub,
                    ufn,
                    key="download_saved_um",
                )
    
    with tab2:
        ctrl_cat = st.selectbox(
            "Kontrol √ľ√ß√ľn baza sah…ôsi:",
            ["Horeca", "Dark Kitchen"],
            key="tab2_cat",
        )
        f_orig = st.file_uploader("1. Orijinal √áek", type=["xlsx"], key="ko")
        f_bot = st.file_uploader("2. Analiz FaylńĪ", type=["xlsx"], key="kb")
        if f_orig and f_bot and st.button("ūüĒć Yoxla"):
            df_o, df_b = pd.read_excel(f_orig), pd.read_excel(f_bot)
            df_o = standardize_columns(df_o, chek_fayli=True)
            df_b = standardize_columns(df_b, chek_fayli=False)
            db = get_db(curr, ctrl_cat)
            if db is not None:
                db = standardize_columns(db, chek_fayli=False)
                db["ad"] = db["ad"].astype(str).str.strip()
                db["id"] = pd.to_numeric(db["id"], errors="coerce")
                db = db.dropna(subset=["id", "ad"])
                db["id"] = db["id"].astype(int)
                db = db.drop_duplicates(subset=["ad"], keep="first")
                if "id" not in df_b.columns:
                    st.markdown("**Problem:** Analiz faylńĪnda `ID` / `id` s√ľtunu tapńĪlmadńĪ.")
                    st.stop()
                bot_ids = set(df_b["id"].astype(int).tolist())
                missing = []
                for _, row in df_o.iterrows():
                    name = str(row.get("ad", ""))
                    p_name, _, _ = apply_special_logic(name, 1, curr)
                    m_name, _ = get_best_match(
                        p_name,
                        _clean_ad_choices(db),
                        threshold=74,
                        safe_mode=True,
                    )
                    tid, _ = (None, None)
                    if m_name:
                        tid, _ = _resolve_id_for_product(db, m_name)
                    if tid is None:
                        tid, _ = _resolve_id_for_product(db, p_name)
                    if tid is not None:
                        if tid not in bot_ids:
                            missing.append(name)
                    else:
                        missing.append(f"{name} (Bazada yoxdur)")
                st.table(pd.DataFrame(missing, columns=["TapńĪlmayanlar"]))
            else:
                st.markdown("**Problem:** Uyńüun ana baza tapńĪlmadńĪ.")


def _render_site_info_dialog_body() -> None:
    """√úmumi t…ôlimat ‚ÄĒ yeni istifad…ô√ßi √ľ√ß√ľn strukturlu m…ôlumat."""
    st.caption(
        "Bu p…ônc…ôr…ôni ist…ônil…ôn vaxt a√ßmaq √ľ√ß√ľn s…ôhif…ônin **…ôn aŇüańüńĪsńĪnda, sańü t…ôr…ôfd…ô** "
        "**‚ĄĻÔłŹ M…ôlumat** d√ľym…ôsini sńĪxńĪn. Qaydalar t…ôtbiqd…ô d…ôyiŇüiklik olduqca bu b√∂lm…ôd…ô yenil…ônir."
    )
    st.markdown(
        "Bu s…ôhif…ô **iki …ôsas Ňü√∂b…ôd…ôn** ibar…ôtdir. Sol paneld…ô **Ňě√∂b…ô** se√ßimin…ô g√∂r…ô "
        "…ôsas iŇü sah…ôsi d…ôyiŇüir. AŇüańüńĪdakńĪ b√∂lm…ôl…ôr h…ôr addńĪmńĪ izah edir."
    )
    st.divider()

    st.subheader("1. Ňě√∂b…ôl…ôr (sol panel)")
    st.markdown(
        """
- **Restoran** ‚ÄĒ sklad **√ß…ôk** faylńĪnńĪn Clopos **ana baza** il…ô uyńüunlaŇüdńĪrńĪlmasńĪ (Analiz) v…ô n…ôtic…ônin yoxlanmasńĪ (Kontrol).
- **ńįnventarizasiya** ‚ÄĒ Clopos-dan g√∂t√ľr√ľlm√ľŇü **h…ôft…ôlik / ay** inventar Excel √ßńĪxńĪŇülarńĪnńĪn eyni qaydalarla emalńĪ (s√ľtun t…ômizliyi, sńĪra, filtr).

Sol paneld…ô **Restoran** aktiv olanda h…ômin restoranńĪn d√ľym…ôsini se√ßin; **ńįnventarizasiya** aktiv olanda is…ô …ôvv…ôl inventar restoranńĪnńĪ se√ßin, sonra fayllarńĪ …ôsas sah…ôd…ôki **1Week ‚Ä¶ MONTH** p…ônc…ôr…ôl…ôrin…ô y√ľkl…ôyin.
        """
    )

    st.subheader("2. Restoran Ňü√∂b…ôsi ‚ÄĒ m…ôlumat m…ônb…ôyi")
    st.markdown(
        """
- **Ana baza** fayllarńĪ **GitHub** m…ônb…ôsind…ôn avtomatik oxunur (…ôl il…ô y√ľkl…ôm…ô t…ôl…ôb olunmur).
- Se√ßdiyiniz restoran v…ô sah…ô (**Horeca** / **Dark Kitchen**) b√ľt√ľn Analiz v…ô Kontrol …ôm…ôliyyatlarńĪ √ľ√ß√ľn …ôsasdńĪr.
        """
    )

    st.subheader("3. ANALńįZ tabńĪ (Restoran)")
    st.markdown(
        """
**M…ôqs…ôd:** √áekd…ôki m…ôhsul **Ad** s√ľtunu il…ô bazadakńĪ **eyni m…ôhsul adńĪ** √ľz-√ľz…ô g…ôtirilir; export faylńĪnda **ID** yalnńĪz bazadan g√∂t√ľr√ľl√ľr.

| AnlayńĪŇü | QńĪsa izah |
|--------|------------|
| **QUANTITY** | √áekd…ôki miqdar (x√ľsusi qayda varsa √ßevrilmiŇü miqdar). |
| **COST** | √áekd…ôki vahid qiym…ôt v…ô miqdar …ôsasńĪnda hesab; qayda **faktoru** varsa (m…ôs. 1 paket = 5 kq) vahid qiym…ôt baza vahidin…ô (m…ôs. 1 kq) √ßevrilir. |
| **T…ôhl√ľk…ôsiz rejim** | YalnńĪz kifay…ôt q…ôd…ôr aydńĪn uyńüunluq q…ôbul edilir; qalńĪqlar **TapńĪlmayanlar** v…ôr…ôqind…ô …ôl il…ô tamamlama √ľ√ß√ľn qalńĪr. |
| **Dark Kitchen** | Ana baza h…ôm `_dk`, h…ôm `_horeca` faylńĪndan oxunub birl…ôŇüdirilir. |
| **Horeca** | YalnńĪz `_horeca` baza faylńĪ istifad…ô olunur. |

∆Źlav…ô se√ßiml…ôr: **Agressiv uyńüunluq** (daha √ßox avtomatik s…ôtir, daha √ßox s…ôhv riski), **uyńüunluq h…ôddi** (%), sklad √ßeki y√ľkl…ôm…ô v…ô **BaŇülat** d√ľym…ôsi.
        """
    )

    st.subheader("4. KONTROL tabńĪ (Restoran)")
    st.markdown(
        """
**M…ôqs…ôd:** **Orijinal √ßek** il…ô **Analiz n…ôtic…ôsi** (ixrac faylńĪ) m√ľqayis…ô olunur.

1. Kontrol √ľ√ß√ľn sah…ô se√ßin (**Horeca** / **Dark Kitchen**).
2. Orijinal √ßek v…ô Analiz faylńĪnńĪ y√ľkl…ôyin.
3. **Yoxla** d√ľym…ôsi il…ô √ßekd…ô g√∂zl…ônil…ôn ID-l…ôrin analiz faylńĪnda olub-olmadńĪńüńĪ yoxlanńĪr; √ßatńĪŇümayanlar c…ôdv…ôld…ô g√∂st…ôrilir.
        """
    )

    st.subheader("5. ńįnventarizasiya Ňü√∂b…ôsi ‚ÄĒ addńĪm-addńĪm")
    st.markdown(
        """
**Fayl y√ľkl…ôm…ô**

- **1Week, 2Week, 3Week, 4Week, MONTH** ‚ÄĒ h…ôr biri √ľ√ß√ľn ayrńĪca **.xlsx** y√ľkl…ôy…ô bil…ôrsiniz.
- Eyni p…ônc…ôr…ôd…ô yeni fayl se√ßdikd…ô **orijinal** n√ľsx…ô yenil…ônir.
- Fayllar restoran √ľzr…ô serverd…ô saxlanńĪlńĪr; s…ôhif…ô yenil…ôns…ô d…ô yenid…ôn g√∂r√ľn√ľr.

**Birinci endirm…ô ‚ÄĒ Orijinal**

- Y√ľkl…ôdiyiniz faylńĪn d…ôyiŇüdirilm…ômiŇü sur…ôti.

**ńįkinci endirm…ô ‚ÄĒ Kateqoriya + sńĪra + filtr** (t…ôk emal faylńĪ)

1. **Kateqoriya (s√ľtun t…ômizliyi v…ô sńĪra)**  
   - Orijinal c…ôdv…ôld…ô Excel s√ľtunlarńĪ **A, B, D, F, K, L, O, P, Q, U** (m√∂vqey…ô g√∂r…ô) silinir.  
   - Qalan c…ôdv…ôld…ô **birinci s√ľtun** (Ňüablonda **Kateqoriya**) √ľzr…ô **A‚ÜíZ** m…ôtn sńĪrasńĪ il…ô sńĪralanńĪr.

2. **Filtr** (sńĪradan d…ôrhal sonra, parametrl…ôr expandable paneld…ô)  
   - ńįst…ôy…ô g√∂r…ô: boŇü **Kateqoriya** s…ôtirl…ôri silinir.  
   - ńįst…ôy…ô g√∂r…ô: boŇü **M…ôhsul** s…ôtirl…ôri silinir.  
   - ńįst…ôy…ô g√∂r…ô: **F…ôrqin d…ôy…ôri** s√ľtununda (**baŇülńĪq** il…ô tapńĪlńĪr; yoxdursa, emaldan sonra **J** m√∂vqesi g√∂t√ľr√ľl√ľr) d…ôy…ôri **sńĪx** olaraq **-10 il…ô 10 arasńĪnda** olan s…ôtirl…ôr √ßńĪxarńĪlńĪr (**-10** v…ô **10** saxlanńĪlńĪr).  
   - R…ôq…ôml…ôr Az…ôrbaycan Excel formatńĪnda ola bil…ôr (m…ôs. onluq **verg√ľl**).

**X…ôtalar**

- ∆Źg…ôr emal alńĪnmazsa, p…ônc…ôr…ôd…ô **‚ö†** il…ô qńĪsa s…ôb…ôb g√∂st…ôrilir; fayl strukturunu v…ô s√ľtun adlarńĪnńĪ yoxlayńĪn.
        """
    )

    st.divider()
    st.caption(
        "∆Źlav…ô sual v…ô ya yeni qayda t…ôklifi √ľ√ß√ľn inkiŇüaf komandasńĪ il…ô …ôlaq…ô saxlayńĪn."
    )


@st.dialog("‚ĄĻÔłŹ Sayt haqqńĪnda ‚ÄĒ qaydalar v…ô addńĪmlar", width="large")
def _open_site_info_dialog() -> None:
    _render_site_info_dialog_body()


# --- SńįDEBAR ---
if "panel_branch" not in st.session_state:
    st.session_state.panel_branch = "restoran"

res_options = discover_restaurants()
if st.session_state.selected_res not in res_options:
    st.session_state.selected_res = res_options[0]
if "selected_inv_res" not in st.session_state:
    st.session_state.selected_inv_res = st.session_state.selected_res
if st.session_state.selected_inv_res not in res_options:
    st.session_state.selected_inv_res = res_options[0]

st.sidebar.markdown("#### Ňě√∂b…ô")
st.sidebar.radio(
    "Ňě√∂b…ô",
    ["restoran", "inventar"],
    horizontal=True,
    key="panel_branch",
    label_visibility="collapsed",
    format_func=lambda x: "ūüŹĘ Restoran" if x == "restoran" else "ūüď¶ ńįnventarizasiya",
)

if st.session_state.panel_branch == "restoran":
    st.sidebar.markdown("##### Restoran se√ßimi")
    for res_opt in res_options:
        label = f"{res_opt} ‚úÖ" if st.session_state.selected_res == res_opt else res_opt
        if st.sidebar.button(label, key=f"btn_{res_opt}", use_container_width=True):
            st.session_state.selected_res = res_opt
            st.rerun()
elif st.session_state.panel_branch == "inventar":
    st.sidebar.markdown("##### ńįnventarizasiya restoranńĪ")
    st.sidebar.selectbox(
        "Restoran",
        res_options,
        key="selected_inv_res",
        help="Fayllar restoran √ľzr…ô ayrńĪca saxlanńĪlńĪr.",
    )

# --- PANELL∆ŹR ---
if st.session_state.panel_branch == "inventar":
    inv_rest = st.session_state.selected_inv_res
    inv_remote_on = _inv_remote_cfg() is not None
    st.markdown(
        f"<h3 style='text-align: center;'>ūüď¶ ńįnventarizasiya | {inv_rest}</h3>",
        unsafe_allow_html=True,
    )
    if inv_remote_on:
        st.success("Deploy-proof yaddaŇü aktivdir (Supabase remote storage).")
    else:
        st.warning(
            "HazńĪrda lokal yaddaŇü aktivdir. Deploy sonrasńĪ saxlama √ľ√ß√ľn `st.secrets`-d…ô "
            "`SUPABASE_URL` v…ô `SUPABASE_SERVICE_ROLE_KEY` …ôlav…ô edin."
        )
    st.markdown(
        "**1‚Äď4 h…ôft…ô:** y√ľkl…ôn…ôn faylńĪn **orijinalńĪ** restoran √ľzr…ô yadda saxlanńĪlńĪr; **ikinci endirm…ô** ‚ÄĒ …ôvv…ôl "
        "Excel **A,B,D,F,K,L,O,P,Q,U** silinir, **A s√ľtunu** (Kateqoriya) √ľzr…ô A‚ÜíZ sńĪra, sonra **filtr**: boŇü s…ôtirl…ôr "
        "v…ô **F…ôrqin d…ôy…ôri** (J): **-10-dan b√∂y√ľk v…ô 10-dan ki√ßik** sńĪx intervaldakńĪ r…ôq…ôml…ôr silinir (**-10** v…ô **10** saxlanńĪr) (t…ôk `.xlsx`). "
        "**MONTH** sonrakńĪ yekunlaŇüdńĪrma √ľ√ß√ľn saxlanńĪr."
    )
    with st.expander("ūüĒé Filtr parametrl…ôri (ikinci endirm…ôd…ô)", expanded=False):
        st.caption(
            "SńĪra (A‚ÜíZ) t…ôtbiq olunduqdan sonra: boŇü s…ôtirl…ôr v…ô **F…ôrqin d…ôy…ôri** √ľzr…ô sńĪra filtri."
        )
        st.checkbox(
            "BoŇü **Kateqoriya** s…ôtirl…ôrini sil",
            value=True,
            key="inv_filter_drop_empty_kat",
        )
        st.checkbox(
            "BoŇü **M…ôhsul** s…ôtirl…ôrini sil",
            value=True,
            key="inv_filter_drop_empty_mah",
        )
        st.checkbox(
            "**F…ôrqin d…ôy…ôri** (J): -10 il…ô 10 arasńĪ **sńĪx** intervaldakńĪ s…ôtirl…ôri sil (-10 v…ô 10 saxla)",
            value=True,
            key="inv_filter_exclude_farqin_mid",
        )

    inv_slots = [
        ("1Week", "inv_week1"),
        ("2Week", "inv_week2"),
        ("3Week", "inv_week3"),
        ("4Week", "inv_week4"),
        ("MONTH", "inv_month"),
    ]
    loaded_weeks = _inv_count_saved_weeks(inv_rest)
    if loaded_weeks == 4:
        st.success("1-4 h…ôft…ô fayllarńĪ bu restoran √ľ√ß√ľn saxlanńĪlńĪb. MONTH yekununa ke√ß…ô bil…ôrsiniz.")
    else:
        st.info(f"Bu restoran √ľ√ß√ľn saxlanmńĪŇü h…ôft…ô fayllarńĪ: {loaded_weeks}/4")

    inv_cols = st.columns(5)
    for inv_col, (inv_label, inv_key) in zip(inv_cols, inv_slots):
        with inv_col:
            with st.container(border=True):
                uploader_key = f"{inv_key}_{inv_rest}_uploader"
                up = st.file_uploader(inv_label, type=["xlsx"], key=uploader_key)
                if up is not None:
                    _inv_store_file(inv_rest, inv_key, up.name, up.getvalue())
                saved = _inv_get_file(inv_rest, inv_key)
                if not saved:
                    continue
                orig = saved["bytes"]
                stem = os.path.splitext(saved["name"])[0][:80]
                st.caption(f"SaxlanńĪb: `{saved['name']}`")
                if st.button("ūüóĎÔłŹ Sil", key=f"inv_clear_{inv_key}_{inv_rest}", use_container_width=True):
                    _inv_delete_file(inv_rest, inv_key)
                    st.session_state.pop(uploader_key, None)
                    st.rerun()
                st.download_button(
                    "ūüď• Orijinal",
                    data=orig,
                    file_name=f"{inv_label}_{stem}_orijinal.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"inv_dl_orig_{inv_key}_{inv_rest}",
                    use_container_width=True,
                )
                _inv_opts = InventoryFilterOptions(
                    drop_empty_kateqoriya=st.session_state.get(
                        "inv_filter_drop_empty_kat", True
                    ),
                    drop_empty_mahsul=st.session_state.get(
                        "inv_filter_drop_empty_mah", True
                    ),
                    exclude_farqin_open_interval_neg10_pos10=st.session_state.get(
                        "inv_filter_exclude_farqin_mid", True
                    ),
                )
                proc_emal, err_emal = process_inventory_emal_pipeline(
                    orig, filter_options=_inv_opts
                )
                if err_emal:
                    st.caption(f"‚ö† {err_emal}")
                else:
                    st.download_button(
                        "ūüď• Kateqoriya + sńĪra + filtr",
                        data=proc_emal,
                        file_name=f"{inv_label}_{stem}_kateqoriya_emal.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"inv_dl_proc_{inv_key}_{inv_rest}",
                        use_container_width=True,
                    )
elif st.session_state.panel_branch == "restoran":
    _render_restoran_online_panel()

st.markdown(
    '<div style="height:2rem" aria-hidden="true"></div>',
    unsafe_allow_html=True,
)
_info_l, _info_r = st.columns([3, 1])
with _info_r:
    if st.button(
        "‚ĄĻÔłŹ M…ôlumat",
        key="site_info_fab_btn",
        help="Sayt Ňü√∂b…ôl…ôri, Analiz, Kontrol v…ô ńįnventarizasiya qaydalarńĪ",
        use_container_width=True,
        type="secondary",
    ):
        _open_site_info_dialog()
